from flask import (
    Flask,
    render_template,
    session,
    request,
    redirect,
    flash,
    url_for,
    send_file
)

from dotenv import load_dotenv

from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename

from functools import wraps

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import psycopg2
import os

# =========================
# ENV
# =========================

load_dotenv()


# =========================
# FLASK
# =========================

app = Flask(__name__)

app.secret_key = os.getenv("SECRET_KEY", "dev-secret")


# =========================
# UPLOAD FOLDER
# =========================

UPLOAD_FOLDER = os.path.join("static", "uploads")

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# =========================
# DATABASE
# =========================

conn = psycopg2.connect(

    host=os.getenv("DB_HOST"),

    database=os.getenv("DB_NAME"),

    user=os.getenv("DB_USER"),

    password=os.getenv("DB_PASSWORD"),

    port=os.getenv("DB_PORT")

)

conn.autocommit = True

# =========================
# ROLE REQUIRED
# =========================

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):

            if "user_id" not in session:
                return redirect(url_for("login"))

            user_role = (session.get("role") or "").lower()
            allowed = [r.lower() for r in roles]

            if user_role not in allowed:
                flash(
                    "You don't have permission to access this page.",
                    "warning"
                )
                return render_template("errors/403.html"), 403

            return f(*args, **kwargs)

        return wrapper
    return decorator


# =========================
# LOGIN REQUIRED
# =========================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):

        if "user_id" not in session:
            return redirect(url_for("login"))

        return f(*args, **kwargs)

    return decorated_function


# =========================
# LOGIN
# =========================

@app.route("/login", methods=["GET", "POST"])
def login():

    if session.get("user_id"):
        return redirect("/")

    if request.method == "POST":

        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")



        # =========================
        # DATABASE LOGIN
        # =========================

        cur = conn.cursor()

        try:

            cur.execute("""
                SELECT id, username, password, role
                FROM inventory.users
                WHERE LOWER(username) = %s
            """, (username,))

            user = cur.fetchone()

        finally:
            cur.close()

        if user:

            db_password = user[2]

            try:
                password_ok = check_password_hash(db_password, password)
            except Exception:
                password_ok = (db_password == password)

            if password_ok:

                session.clear()

                session["user_id"] = user[0]
                session["username"] = user[1]
                session["role"] = (user[3] or "").lower()

                flash("Welcome back!", "success")
                return redirect("/")

        flash("Invalid username or password.", "danger")

    return render_template("login.html")

# =========================
# LOGOUT
# =========================

@app.route("/logout")
@login_required
def logout():

    session.clear()

    flash(
        "Logged out successfully.",
        "success"
    )

    return redirect("/login")


# =========================
# DASHBOARD
# =========================

@app.route("/")
@login_required
def index():

    cur = conn.cursor()

    try:

        # =========================
        # PRODUCTS
        # =========================

        cur.execute("""
            SELECT
                id,
                product_name,
                price_numeric,
                image,
                quantity,
                category
            FROM inventory.products
            ORDER BY id
        """)

        products = cur.fetchall()


        # =========================
        # DASHBOARD CARDS
        # =========================

        total_products = len(products)

        total_stock = sum(
            product[4] or 0
            for product in products
        )


        inventory_value = sum(
            float(product[2] or 0) * (product[4] or 0)
            for product in products
        )


        low_stock = sum(
            1
            for product in products
            if (product[4] or 0) <= 5
        )


        # =========================
        # PIE CHART
        # Products per Category
        # =========================

        cur.execute("""
            SELECT
                category,
                COUNT(*)
            FROM inventory.products
            GROUP BY category
            ORDER BY category
        """)

        category_data = cur.fetchall()



        # =========================
        # BAR CHART
        # Stock per Product
        # =========================

        cur.execute("""
            SELECT
                product_name,
                quantity
            FROM inventory.products
            ORDER BY product_name
        """)

        stock_data = cur.fetchall()



        # =========================
        # CHART DATA
        # =========================

        category_labels = [
            row[0] or "No Category"
            for row in category_data
        ]


        category_values = [
            row[1]
            for row in category_data
        ]


        stock_labels = [
            row[0]
            for row in stock_data
        ]


        stock_values = [
            row[1] or 0
            for row in stock_data
        ]



        # =========================
        # RENDER
        # =========================

        return render_template(
            "index.html",

            products=products,

            total_products=total_products,
            total_stock=total_stock,
            inventory_value=inventory_value,
            low_stock=low_stock,

            category_labels=category_labels,
            category_values=category_values,

            stock_labels=stock_labels,
            stock_values=stock_values
        )


    finally:

        cur.close()


# =========================
# ADD PRODUCT
# =========================

@app.route("/add", methods=["GET", "POST"])
@login_required
@role_required("admin")
def add_product():

    

    if request.method == "POST":

        name = request.form["product_name"]
        category = request.form["category"]
        price = request.form["price"]
        quantity = request.form["quantity"]


        # =========================
        # IMAGE UPLOAD
        # =========================

        image = request.files.get("image")


        if image and image.filename != "":

            filename = secure_filename(image.filename)

            image.save(
                os.path.join(
                    app.config["UPLOAD_FOLDER"],
                    filename
                )
            )

            image_filename = filename

        else:

            image_filename = None



        cur = conn.cursor()


        try:


            # =========================
            # INSERT PRODUCT
            # =========================

            cur.execute("""
                INSERT INTO inventory.products
                (
                    product_name,
                    price_numeric,
                    image,
                    quantity,
                    category
                )

                VALUES (%s,%s,%s,%s,%s)

                RETURNING id

            """,
            (
                name,
                price,
                image_filename,
                quantity,
                category
            ))


            product_id = cur.fetchone()[0]



            # =========================
            # SAVE TRANSACTION
            # =========================

            cur.execute("""
                INSERT INTO inventory.inventory_transactions
                (
                    product_id,
                    transaction_type,
                    quantity,
                    user_id
                )

                VALUES
                (
                    %s,
                    'ADD PRODUCT',
                    %s,
                    %s
                )

            """,
            (
                product_id,
                quantity,
                session["user_id"]
            ))



            flash(
                "Product added successfully!",
                "success"
            )


        finally:

            cur.close()


        return redirect("/")


    return render_template("add.html")


# =========================
# EDIT PRODUCT
# =========================

@app.route("/edit/<int:id>", methods=["GET", "POST"])
@login_required
@role_required("admin")
def edit_product(id):

    cur = conn.cursor()

    try:


        # =========================
        # UPDATE PRODUCT
        # =========================

        if request.method == "POST":

            name = request.form.get("product_name", "").strip()
            category = request.form.get("category", "").strip()
            price = request.form.get("price", 0)
            quantity = request.form.get("quantity", 0)


            cur.execute("""
                UPDATE inventory.products
                SET
                    product_name = %s,
                    price_numeric = %s,
                    quantity = %s,
                    category = %s

                WHERE id = %s

            """,
            (
                name,
                price,
                quantity,
                category,
                id
            ))


            flash(
                "Product updated successfully!",
                "success"
            )


            return redirect("/")



        # =========================
        # LOAD PRODUCT
        # =========================

        cur.execute("""
            SELECT

                id,
                product_name,
                price_numeric,
                quantity,
                category

            FROM inventory.products

            WHERE id = %s

        """, (id,))


        product = cur.fetchone()



        if not product:


            flash(
                "Product not found.",
                "danger"
            )


            return redirect("/")



        return render_template(
            "edit.html",
            product=product
        )



    finally:

        cur.close()


# =========================
# DELETE PRODUCT
# =========================

@app.route("/delete/<int:id>", methods=["POST"])
@login_required
@role_required("admin")
def delete_product(id):

    cur = conn.cursor()

    try:

        # =========================
        # CHECK PRODUCT
        # =========================

        cur.execute("""
            SELECT id
            FROM inventory.products
            WHERE id = %s
        """, (id,))

        product = cur.fetchone()

        if not product:

            flash(
                "Product not found.",
                "danger"
            )

            return redirect("/")

        # =========================
        # DELETE TRANSACTION HISTORY
        # =========================

        cur.execute("""
            DELETE FROM inventory.inventory_transactions
            WHERE product_id = %s
        """, (id,))

        # =========================
        # DELETE PRODUCT
        # =========================

        cur.execute("""
            DELETE FROM inventory.products
            WHERE id = %s
        """, (id,))

        flash(
            "Product deleted successfully!",
            "success"
        )

    except psycopg2.Error as e:

        flash(
            f"Database Error: {e.pgerror or str(e)}",
            "danger"
        )

    finally:

        cur.close()

    return redirect("/")


# =========================
# STOCK IN
# =========================

@app.route("/stock-in/<int:id>", methods=["POST"])
@login_required
@role_required("admin")
def stock_in(id):

    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT id
            FROM inventory.products
            WHERE id = %s
        """, (id,))

        product = cur.fetchone()

        if not product:
            flash("Product not found.", "danger")
            return redirect("/")

        cur.execute("""
            UPDATE inventory.products
            SET quantity = quantity + 1
            WHERE id = %s
        """, (id,))

        cur.execute("""
            INSERT INTO inventory.inventory_transactions
            (
                product_id,
                transaction_type,
                quantity,
                user_id
            )
            VALUES (%s,'STOCK IN',1,%s)
        """, (id, session["user_id"]))

        flash("Stock added successfully!", "success")

    except Exception as e:
        print("=" * 60)
        print(e)
        print("=" * 60)
        raise

    finally:
        cur.close()

    return redirect("/")

# =========================
# STOCK OUT
# =========================

@app.route("/stock-out/<int:id>", methods=["POST"])
@login_required
@role_required("admin")
def stock_out(id):

    cur = conn.cursor()

    try:


        # Check product

        cur.execute("""
            SELECT
                id,
                quantity

            FROM inventory.products

            WHERE id = %s

        """, (id,))



        product = cur.fetchone()



        if not product:


            flash(
                "Product not found.",
                "danger"
            )

            return redirect("/")



        # Check stock

        if product[1] <= 0:


            flash(
                "No stock available to deduct.",
                "warning"
            )

            return redirect("/")



        # Decrease stock

        cur.execute("""
            UPDATE inventory.products

            SET quantity = quantity - 1

            WHERE id = %s

        """, (id,))



        # Save transaction

        cur.execute("""
            INSERT INTO inventory.inventory_transactions

            (
                product_id,
                transaction_type,
                quantity,
                user_id
            )

            VALUES

            (
                %s,
                'STOCK OUT',
                1,
                %s
            )

        """,
        (
            id,
            session["user_id"]
        ))



        flash(
            "Stock deducted successfully!",
            "success"
        )



    finally:

        cur.close()



    return redirect("/")


# =========================
# EXPORT TRANSACTIONS EXCEL
# =========================

@app.route("/transactions/export")
@login_required
@role_required("admin")
def export_transactions_excel():

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["ID", "Product", "Type", "Quantity", "Date", "User"]
    ws.append(headers)

    # Style header
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    t.id,
                    p.product_name,
                    t.transaction_type,
                    t.quantity,
                    t.transaction_date,
                    COALESCE(u.username, 'Demo User')
                FROM inventory.inventory_transactions t
                JOIN inventory.products p ON p.id = t.product_id
                LEFT JOIN inventory.users u ON u.id = t.user_id
                ORDER BY t.transaction_date DESC
            """)

            rows = cur.fetchall()

        for r in rows:
            ws.append([
                r[0],
                r[1],
                r[2],
                r[3],
                r[4].strftime("%Y-%m-%d %H:%M:%S") if r[4] else "",
                r[5]
            ])

    except Exception as e:
        flash(f"Export error: {str(e)}", "danger")
        return redirect(url_for("transactions"))

    # Auto width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="transactions.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
# =========================
# EXPORT PRODUCTS EXCEL
# =========================

@app.route("/products/export")
@login_required
@role_required("admin")
def export_products_excel():

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ["ID", "Product", "Price", "Quantity", "Category"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="2E7D32")
        cell.alignment = Alignment(horizontal="center")

    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, product_name, price_numeric, quantity, category
                FROM inventory.products
                ORDER BY id
            """)

            rows = cur.fetchall()

        for r in rows:
            ws.append([
                r[0],
                r[1],
                r[2],
                r[3],
                r[4]
            ])

    except Exception as e:
        flash(f"Export error: {str(e)}", "danger")
        return redirect(url_for("index"))

    # Auto width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="products.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =========================
# TRANSACTIONS
# =========================

@app.route("/transactions")
@login_required
@role_required("admin", "staff", "demo")
def transactions():

    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                t.id,
                p.product_name,
                t.transaction_type,
                t.quantity,
                t.transaction_date,
                COALESCE(u.username, 'Demo User') AS username
            FROM inventory.inventory_transactions t
            JOIN inventory.products p ON p.id = t.product_id
            LEFT JOIN inventory.users u ON u.id = t.user_id
            ORDER BY t.transaction_date DESC
        """)

        rows = cur.fetchall()

    finally:
        cur.close()

    return render_template(
        "transactions.html",
        transactions=rows
    )

# =========================
# REPORTS
# =========================

@app.route("/reports")
@login_required
def reports():

    cur = conn.cursor()


    try:


        # =========================
        # PRODUCTS
        # =========================

        cur.execute("""
            SELECT

                id,
                product_name,
                price_numeric,
                image,
                quantity,
                category


            FROM inventory.products


            ORDER BY id

        """)


        products = cur.fetchall()



        # =========================
        # SUMMARY
        # =========================


        total_products = len(products)


        total_stock = sum(
            p[4] or 0
            for p in products
        )


        inventory_value = sum(
            float(p[2] or 0) * (p[4] or 0)

            for p in products
        )


        low_stock = sum(
            1

            for p in products

            if (p[4] or 0) <= 5

        )



        # =========================
        # PIE CHART
        # =========================

        cur.execute("""
            SELECT

                category,
                COUNT(*)


            FROM inventory.products


            GROUP BY category


            ORDER BY category

        """)


        category_data = cur.fetchall()



        # =========================
        # BAR CHART
        # =========================

        cur.execute("""
            SELECT

                product_name,
                quantity


            FROM inventory.products


            ORDER BY product_name

        """)


        stock_data = cur.fetchall()



        category_labels = [
            r[0] or "No Category"

            for r in category_data
        ]


        category_values = [
            r[1]

            for r in category_data
        ]



        stock_labels = [
            r[0]

            for r in stock_data
        ]


        stock_values = [
            r[1] or 0

            for r in stock_data
        ]



        return render_template(

            "reports.html",


            products=products,


            total_products=total_products,
            total_stock=total_stock,
            inventory_value=inventory_value,
            low_stock=low_stock,


            category_labels=category_labels,
            category_values=category_values,


            stock_labels=stock_labels,
            stock_values=stock_values

        )



    finally:

        cur.close()

# =========================
# ERROR PAGES
# =========================

@app.errorhandler(404)
def not_found(error):
    return render_template("errors/404.html"), 404


@app.errorhandler(403)
def forbidden(error):
    return render_template("errors/403.html"), 403


@app.errorhandler(500)
def server_error(error):
    return render_template("errors/500.html"), 500        

# =========================
# RUN
# =========================

if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=5000,
        debug=True
    )